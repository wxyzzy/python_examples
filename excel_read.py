# Read from excel workbook
# This reads what was written by excel_write.py
# It expects a table written in sheet 'mult'


import openpyxl

filename = 'excel_test_1.xls'

def cell(x,y):
    # convert coordinate to excel address
    a = chr(ord('a') + x - 1)
    b = str(y)
    return a + b
    
def read_from_excel():
    try:
        wb = openpyxl.load_workbook(filename)
    except Exception as e:
        print(e)
        print('Workbook ' + filename + ' not found or could not be opened.')
        return
    if 'mult' not in wb.sheetnames:
        print('Sheet not found')
        return
    ws = wb['mult']
    lst = [[x.value for x in row] for row in ws.iter_rows()]
    print(lst)
    print('Example value: {}'.format(ws[cell(9,9)].value))

read_from_excel()
